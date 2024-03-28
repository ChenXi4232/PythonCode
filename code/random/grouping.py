import pandas as pd
from pandas.core.frame import DataFrame
import random
import openpyxl


def sort_by_subject():  # 根据所选科目分为物理组和历史组
    table = openpyxl.load_workbook('D:\\Py_program\\code\\random\\data1.xlsx')
    sheet = table['Sheet1']
    new_table = openpyxl.Workbook()
    sheet1 = new_table.active
    sheet1.title = '物理组'
    df = pd.read_excel("D:\\Py_program\\code\\random\\data1.xlsx")
    title = df.columns
    sheet1.append(list(title))
    for row in sheet.values:
        if type(row[6]) == int:
            sheet1 = new_table['物理组']
            sheet1.append(list(row))
        else:
            try:
                sheet1 = new_table['历史组']
            except KeyError:
                sheet1 = new_table.create_sheet('历史组')
            finally:
                sheet1 = new_table['历史组']
                sheet1.append(list(row))
    new_table.save('按照学科分类.xlsx')


def physics(physics_groups):  # 对物理组分组
    table = openpyxl.load_workbook('按照学科分类.xlsx')
    sheet = table['物理组']
    physics_male = []
    physics_female = []
    for row in sheet.values:
        if (row[2] == '男'):
            physics_male.append(row[1])
        if (row[2] == '女'):
            physics_female.append(row[1])

    random.shuffle(physics_male)
    random.shuffle(physics_female)

    group_size = 5
    group_num = int((len(physics_female) + len(physics_male)) / group_size)
    # 平衡每一组的男女比例
    while len(physics_groups) < group_num and (len(physics_male) > 0
                                               and len(physics_female) > 0):
        num_male = int(
            len(physics_male) / (len(physics_female) + len(physics_male)) *
            group_size)
        num_female = group_size - num_male
        males_in_group = [
            physics_male.pop(0) for n in range(num_male)
            if len(physics_male) > 0
        ]
        females_in_group = [
            physics_female.pop(0) for n in range(num_female)
            if len(physics_female) > 0
        ]
        group = males_in_group + females_in_group
        random.shuffle(group)
        physics_groups.append(group)
    # 添加组名
    for n in range(len(physics_groups)):
        physics_groups[n].insert(0, '第%d组' % (n + 1))


def history(history_groups):  # 对历史组分组
    table = openpyxl.load_workbook('按照学科分类.xlsx')
    sheet = table['历史组']
    history_male = []
    history_female = []
    for row in sheet.values:
        if (row[2] == '男'):
            history_male.append(row[1])
        if (row[2] == '女'):
            history_female.append(row[1])

    random.shuffle(history_male)
    random.shuffle(history_female)

    group_size = 5
    group_num = int((len(history_female) + len(history_male)) / group_size)
    # 平衡每一组的男女比例
    while len(history_groups) < group_num and (len(history_male) > 0
                                               and len(history_female) > 0):
        num_male = int(
            len(history_male) / (len(history_female) + len(history_male)) *
            group_size)
        num_female = group_size - num_male
        males_in_group = [
            history_male.pop(0) for n in range(num_male)
            if len(history_male) > 0
        ]
        females_in_group = [
            history_female.pop(0) for n in range(num_female)
            if len(history_female) > 0
        ]
        group = males_in_group + females_in_group
        random.shuffle(group)
        history_groups.append(group)
    # 添加组名
    for n in range(len(history_groups)):
        history_groups[n].insert(0, '第%d组' % (n + 1))


def output_result(physics_groups, history_groups):  # 将分组结果写入excel文件中
    data_physics = DataFrame(physics_groups)
    data_physics = data_physics.T
    df1 = pd.DataFrame(data_physics)
    data_history = DataFrame(history_groups)
    data_history = data_history.T
    df2 = pd.DataFrame(data_history)
    with pd.ExcelWriter('随机分组结果.xlsx') as writer:
        df1.to_excel(writer, sheet_name='物理组', index=False, header=None)
        df2.to_excel(writer, sheet_name='历史组', index=False, header=None)


if __name__ == '__main__':
    sort_by_subject()
    physics_groups = []
    history_groups = []
    physics(physics_groups)
    history(history_groups)
    output_result(physics_groups, history_groups)
